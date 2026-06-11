#!/usr/bin/env python3
"""
nces_school_extract.py  (Dominion VA+NC playbook, Part 13)

Pulls every public school in Virginia + North Carolina from the NCES EDGE
"Public School Locations - Current" FeatureServer and writes
dominion_school_sites.xlsx with the exact column schema the dashboard's
"Upload sites" panel expects:

    NAME, LAT, LON, STATE, COUNTY, NCES_ID, TYPE, ADDRESS, CITY, ZIP

Optional private (PSS) / postsecondary (IPEDS) layers can be added to SOURCES.

Stdlib only except openpyxl:   pip install openpyxl
Run:                            python nces_school_extract.py
"""

import json
import time
import urllib.parse
import urllib.request

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("openpyxl is required:  pip install openpyxl")

# --- sources -----------------------------------------------------------------
# Each source: (url, type_label). Add private/postsecondary layers here later.
SOURCES = [
    ("https://services1.arcgis.com/Ua5sjt3LWTPigjyD/arcgis/rest/services/"
     "Public_School_Locations_Current/FeatureServer/0", "public"),
    # ("<EDGE Private School Locations FeatureServer/0>", "private"),
    # ("<IPEDS / postsecondary FeatureServer/0>",          "postsecondary"),
]

WHERE = "STATE IN ('VA','NC')"   # EDGE current uses a STATE text field ('VA','NC')
PAGE = 2000
OUTFILE = "dominion_school_sites.xlsx"

# Candidate field names per concept (auto-detected against the layer schema).
FIELD_CANDIDATES = {
    "name":   ["NAME", "SCH_NAME", "INST_NAME"],
    "lat":    ["LAT", "LATITUDE", "Y"],
    "lon":    ["LON", "LONGITUDE", "X"],
    "state":  ["STATE", "LSTATE", "STABBR"],
    "county": ["NMCNTY", "CNTY", "COUNTY", "NMCNTY15"],
    "nces":   ["NCESSCH", "NCESID", "UNITID", "PPIN"],
    "street": ["STREET", "ADDRESS", "LSTREET", "STREET1"],
    "city":   ["CITY", "LCITY"],
    "zip":    ["ZIP", "LZIP", "ZIP5"],
}


def fetch_json(url, params):
    qs = urllib.parse.urlencode(params)
    req = urllib.request.Request(url + "?" + qs, headers={"User-Agent": "dominion-playbook/1.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.load(r)


def layer_fields(url):
    meta = fetch_json(url, {"f": "json"})
    return {f["name"].upper(): f["name"] for f in meta.get("fields", [])}


def pick(fields_map, concept):
    for cand in FIELD_CANDIDATES[concept]:
        if cand.upper() in fields_map:
            return fields_map[cand.upper()]
    return None


def extract_source(url, type_label, seen):
    fields_map = layer_fields(url)
    resolved = {k: pick(fields_map, k) for k in FIELD_CANDIDATES}
    rows, offset = [], 0
    while True:
        data = fetch_json(url, {
            "where": WHERE, "outFields": "*", "returnGeometry": "true",
            "outSR": 4326, "f": "json", "resultOffset": offset,
            "resultRecordCount": PAGE, "orderByFields": "OBJECTID",
        })
        feats = data.get("features", [])
        if not feats:
            break
        for f in feats:
            a = f.get("attributes", {})
            g = f.get("geometry", {}) or {}
            lat = a.get(resolved["lat"]) if resolved["lat"] else None
            lon = a.get(resolved["lon"]) if resolved["lon"] else None
            if lat in (None, "") or lon in (None, ""):
                lat, lon = g.get("y"), g.get("x")
            if lat in (None, "") or lon in (None, ""):
                continue
            nces = a.get(resolved["nces"]) if resolved["nces"] else None
            name = a.get(resolved["name"]) if resolved["name"] else ""
            key = nces or (str(name).strip().lower() + f"|{round(float(lat),4)},{round(float(lon),4)}")
            if key in seen:
                continue
            seen.add(key)
            rows.append({
                "NAME": name,
                "LAT": round(float(lat), 6),
                "LON": round(float(lon), 6),
                "STATE": a.get(resolved["state"]) if resolved["state"] else "",
                "COUNTY": a.get(resolved["county"]) if resolved["county"] else "",
                "NCES_ID": nces or "",
                "TYPE": type_label,
                "ADDRESS": a.get(resolved["street"]) if resolved["street"] else "",
                "CITY": a.get(resolved["city"]) if resolved["city"] else "",
                "ZIP": a.get(resolved["zip"]) if resolved["zip"] else "",
            })
        offset += len(feats)
        if len(feats) < PAGE:
            break
        time.sleep(0.2)
    return rows


def main():
    seen, all_rows = set(), []
    for url, label in SOURCES:
        print(f"Querying {label}: {url}")
        rows = extract_source(url, label, seen)
        print(f"  -> {len(rows)} rows")
        all_rows.extend(rows)

    all_rows.sort(key=lambda r: (r["STATE"], str(r["COUNTY"]), str(r["NAME"])))

    cols = ["NAME", "LAT", "LON", "STATE", "COUNTY", "NCES_ID", "TYPE", "ADDRESS", "CITY", "ZIP"]
    wb = Workbook()
    ws = wb.active
    ws.title = "schools"
    ws.append(cols)
    hdr_fill = PatternFill("solid", fgColor="1B2A4A")
    hdr_font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
    for r in all_rows:
        ws.append([r[c] for c in cols])
    ws.freeze_panes = "A2"
    widths = {"NAME": 40, "ADDRESS": 30, "CITY": 18, "COUNTY": 22}
    for i, c in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)
    wb.save(OUTFILE)
    print(f"Wrote {OUTFILE} with {len(all_rows)} schools.")


if __name__ == "__main__":
    main()
